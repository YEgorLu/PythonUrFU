import re
from typing import Tuple, List

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from matplotlib import pyplot as plt
from jinja2 import Environment, FileSystemLoader
import pdfkit


class Report:
    def __init__(self, salary_by_city, count_by_city, salary_by_year, count_by_year, prof_salary_by_year,
                 prof_count_by_year, prof_name: str):
        self.__salary_by_city = salary_by_city
        self.__count_by_city = count_by_city
        self.__salary_by_year = salary_by_year
        self.__count_by_year = count_by_year
        self.__prof_salary_by_year = prof_salary_by_year
        self.__prof_count_by_year = prof_count_by_year
        self.__prof_name = prof_name
        self.__font = Font(bold=True)
        self.__border = Border(
            left=Side(style='thin'),
            top=Side(style='thin'),
            right=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.__fig, self.__axs = plt.subplots(2, 2, layout='tight', figsize=[10, 10])
        self.__first_table_header, self.__first_table = self.__generate_years_table()
        self.__second_table_header, self.__second_table = self.__generate_cities_salary_table()
        self.__third_table_header, self.__third_table = self.__generate_cities_vacancy_table()

    def generate_excel(self):
        wk = Workbook()
        dest_filename = 'report/report.xlsx'
        ws = wk.worksheets[0]
        ws.title = "Статистика по годам"
        ws.append(self.__first_table_header)
        self.__style_header(ws)

        for t in self.__first_table:
            ws.append(t)
        self.__style_cells(ws)

        wk.create_sheet(title="Статистика по городам")
        ws = wk.worksheets[1]
        header = (self.__second_table_header + tuple('') + self.__third_table_header)
        ws.append(header)

        table = [(self.__second_table[i] + tuple(" ") + self.__third_table[i]) for i in range(len(self.__third_table))]
        for t in table:
            ws.append(t)
        self.__style_percent_column(ws, 5)
        self.__style_cells(ws)
        self.__style_header(ws)

        wk.save(filename=dest_filename)

    def generate_image(self):
        self.__generate_salary_diagram()
        self.__generate_vacancy_diagram()
        self.__generate_hor()
        self.__generate_pie()
        plt.savefig('report/graph.png')

    def generate_pdf(self):
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("template/template.html")

        pdf_template = template.render({
            'first': 'graph.png',
            'first_table': self.__first_table,
            'second_table': self.__second_table,
            'first_table_header': self.__first_table_header,
            'second_table_header': self.__second_table_header,
            'third_table_header': self.__third_table_header,
            'third_table': list(
                map(lambda tup: (tup[0], "{:.2f}%".format(tup[1] * 100).replace('.', ',')), self.__third_table)),
            'prof_name': self.__prof_name
        })
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report/report.pdf', configuration=config,
                           options={"enable-local-file-access": ""})

    def __generate_years_table(self) -> (Tuple[str, str, str, str, str], List[Tuple[int, int, int, int, int]]):
        fth = ('Год', 'Средняя зарплата', f'Средняя зарплата - {self.__prof_name}', 'Количество вакансий',
               f'Количество вакансий - {self.__prof_name}')
        years = {k: k for k in self.__count_by_year.keys()}
        ft = [(years[i], self.__salary_by_year[i], self.__prof_salary_by_year[i],
               self.__count_by_year[i], self.__prof_count_by_year[i]) for i in years]
        return fth, ft

    def __generate_cities_salary_table(self) -> (Tuple[str, str], List[Tuple[str, float]]):
        sth = 'Город', 'Уровень зарплат'
        st = [(city, self.__salary_by_city[city]) for city in self.__salary_by_city]
        return sth, st

    def __generate_cities_vacancy_table(self) -> (Tuple[str, str], List[Tuple[str, int]]):
        tth = 'Город', 'Доля вакансий'
        tt = [(city, self.__count_by_city[city]) for city in self.__count_by_city]
        return tth, tt

    def __style_header(self, worksheet: Worksheet) -> None:
        for header in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in header:
                cell.font = self.__font
                cell.border = self.__border

    def __style_cells(self, worksheet: Worksheet) -> None:
        dims = []
        for column in worksheet.columns:
            max_width = 0
            for cell in column:
                if cell.value != " ":
                    cell.border = self.__border
                    if max_width < len(str(cell.value)):
                        max_width = len(str(cell.value))
            dims.append(max_width + 2)
        for i, width in enumerate(dims):
            worksheet.column_dimensions[get_column_letter(i + 1)].width = width

    def __style_percent_column(self, worksheet: Worksheet, col_index: int) -> None:
        for column in worksheet.iter_cols(col_index):
            for cell in column:
                cell.style = 'Percent'
                cell.number_format = '0.00%'

    def __generate_pie(self):
        ax = self.__axs[1, 1]
        data = {'Другие': 1 - sum(self.__count_by_city.values()), **self.__count_by_city}
        ax.pie(list(data.values()), labels=list(data.keys()), textprops={'fontsize': 6})
        ax.set_title('Доля вакансий по городам', fontsize=20)

    def __generate_hor(self):
        ax = self.__axs[1, 0]
        y_pos = range(len(self.__salary_by_city.keys()))
        cities = list(self.__salary_by_city.keys())
        for i, city in enumerate(cities):
            cities[i] = re.sub(r'([- ])', lambda s: '-\n' if s.group(0) == '-' else '\n', city)
        salary = list(self.__salary_by_city.values())
        ax.barh(y_pos, salary)
        ax.set_yticks(y_pos, labels=cities, fontsize=6)
        for tick in ax.get_xticklabels():
            tick.set_fontsize(8)
            tick.ha = 'right'
            tick.va = 'center'
        ax.invert_yaxis()
        ax.set_title('Уровень зарплат по городам', fontsize=20)
        ax.grid(axis='x')

    def __generate_salary_diagram(self):
        ax = self.__axs[0, 0]
        ax.set_title('Уровень зарплат по годам', fontsize=20)
        y_pos = range(len(self.__salary_by_year.keys()))
        years = list(self.__salary_by_year.keys())
        salary_by_year = list(self.__salary_by_year.values())
        prof_salary = list(self.__prof_salary_by_year.values())
        overall = ax.bar(y_pos, salary_by_year, width=0.3)
        prof = ax.bar(list(map(lambda x: x + 0.3, y_pos)), prof_salary, width=0.3)
        ax.set_xticks(list(map(lambda x: x + 0.15, y_pos)), labels=years, rotation="vertical", fontsize=8)
        for tick in ax.get_yticklabels():
            tick.set_fontsize(8)
        ax.legend([overall, prof], [r'средняя з\п', r'з\п программист'], fontsize=8)
        ax.grid(axis='y')

    def __generate_vacancy_diagram(self):
        ax = self.__axs[0, 1]
        y_pos = range(len(self.__count_by_year.keys()))
        years = list(self.__count_by_year.keys())
        salary_by_year = list(self.__count_by_year.values())
        prof_salary = list(self.__prof_count_by_year.values())
        overall = ax.bar(y_pos, salary_by_year, width=0.3)
        prof = ax.bar(list(map(lambda x: x + 0.3, y_pos)), prof_salary, width=0.3)
        ax.set_xticks(list(map(lambda x: x + 0.15, y_pos)), labels=years, rotation="vertical", fontsize=8)
        for tick in ax.get_yticklabels():
            tick.set_fontsize(8)
        ax.legend([overall, prof], ['Количество вакансий', 'Количество ваканисий\n программист'], fontsize=8)
        ax.set_title('Количество вакансий по годам', fontsize=20)
        ax.grid(axis='y')
